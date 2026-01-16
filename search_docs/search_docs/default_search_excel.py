from  search_docs.search_docs import AbstractSearchDocs
import pandas as pd
import openpyxl
import win32com.client
import warnings
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
import os
from bteam_utils import CommonProgress

class DefaultSearchExcel(AbstractSearchDocs):
    """Excelドキュメント検索クラス
    """
    _doc_type: str = 'Excel'                        # ドキュメントタイプをoverride
    _extensions: list = ['.xls', '.xlsx', '.xlsm']  # 対応拡張子リストをoverride

    #
    # constructor/destructor
    #
    def __init__(self, enable_progress: bool = True) -> None:
        """コンストラクタ
        """
        super().__init__(enable_progress)

    def __del__(self) -> None:
        """デストラクタ
        """
        super().__del__()

    #
    # public methods
    #
    def search_element(self, target_path:str) -> bool:
        """ドキュメント要素検索処理
        Args:
            target_path (str): 検索対象パス

        Returns:
            bool: True:成功, False:失敗
        """
        # 初期化
        excel_files = []

        # フォルダ内のexcelファイルリストを取得
        for root, dirs, files in os.walk(target_path):
            for file in files:
                # 拡張子がExcelファイルの場合
                if file.endswith(tuple(self._extensions)):
                    excel_files.append(os.path.join(root, file))

        # excelファイルがある場合はシート名リストを取得
        if len(excel_files) > 0:
            self._search_sheet_list(excel_files)

        # 検索結果に行が存在する場合はTrueを返す
        if not self._pd_element.empty:
            return True
        else:
            return False

    def search_keyword(self, keywords:list, enable_search_shapes: bool = False) -> bool:
        """キーワード検索処理

        このメソッドは、事前にsearch_elementメソッドが実行されていることを前提としています。

        Args:
            keywords (list): 検索キーワード
            enable_search_shapes (bool): 図形内テキスト検索有効フラグ

        Returns:
            bool: True:成功, False:失敗
        """
        # キーワードがない場合は終了
        if not keywords or len(keywords) == 0:
            return False
        # 要素検索結果がない場合は終了
        if self._pd_element is None or self._pd_element.empty:
            return False
        
        # キーワード検索結果を要素検索結果で初期化
        self._pd_keyword = self._pd_element.copy(deep=True)
        # キーワード列を追加
        for keyword in keywords:
            self._pd_keyword[keyword] = None

        # CELL内テキスト検索を実行
        self._search_keyword_cell(keywords)

        # 図形内テキスト検索を実行
        if enable_search_shapes:
            self._search_keyword_shape(keywords)
        
        # キーワード検索結果に行が存在する場合はTrueを返す
        if self._pd_keyword is not None and not self._pd_keyword.empty:
            return True
        else:
            return False
        
    #
    # protected methods
    #
    def _search_sheet_list(self, files:list) -> None:
        """
        シート名リスト取得
        Args:
            files (list): excelファイルのリスト（フルパス）
        """

        # 初期化
        progress_max = len(files)
        rows = []
        workbook = None
        # 進捗表示用フラグを初期化
        progress = CommonProgress(total=progress_max, task_msg=self._doc_type+' Sheets') if self._enable_progress else None

        # ファイルごとにシート名を取得
        for i, file in enumerate(files, 1):
            file_path = os.path.dirname(file)
            file_name = os.path.basename(file)
            try:
                # ブックを開く
                workbook = openpyxl.load_workbook(file, read_only=True, data_only=True)
                sheetnames = workbook.sheetnames
                # ブック名とシート名をリストに登録
                for sheetname in sheetnames:
                    rows.append({
                        'Path':file_path,
                        'Book':file_name,
                        'Sheet':sheetname
                    })
            except :
                # ファイルが開けない場合はエラーメッセージを登録
                rows.append({
                    'Path':file_path,
                    'Book':file_name,
                    'Sheet':"Bad File Error"
                })
            finally:
                # workbookを閉じる
                if workbook is not None:
                    workbook.close()
                    workbook = None

            # 進捗表示
            if progress:
                progress.update(current=i, status_msg=f'Processing: {i}/{progress_max}')

        # 列名を設定（0を空文字に置換）
        self._pd_element = pd.DataFrame(rows, columns=['Path','Book','Sheet']).replace(0, "")

        # 進捗表示(100%)
        if progress:
            progress.update(current=progress_max, status_msg='Completed')
    
    def _search_keyword_cell(self, keywords:list) -> None:
        """キーワード検索処理
        Args:
            keywords (list): 検索キーワードリスト
        """
        # 進捗データ初期化
        progress_cnt = 0
        progress_max = self._pd_keyword.shape[0]
        workbook= None
        current_workbook_path = None
        # 進捗表示用フラグを初期化
        progress = CommonProgress(total=progress_max, task_msg=self._doc_type+' Keyword in Cells') if self._enable_progress else None

        # ブック＋シートでキーワードを検索する
        for output_index, output_row in self._pd_keyword.iterrows():
            # Bad Fileの場合はスキップ
            if output_row['Sheet'] == "Bad File Error":
                progress_cnt += 1
                if progress:
                    progress.update(current=progress_cnt, status_msg=f'Processing: {progress_cnt}/{progress_max}')
                continue
            
            # フルパスを生成してOpen済みか確認
            full_workbook_path = os.path.join(output_row['Path'], output_row['Book'])
            if current_workbook_path != full_workbook_path:
                if workbook is not None:
                    workbook.close()

                current_workbook_path = full_workbook_path
                try:
                    # ブックを開く
                    workbook = openpyxl.load_workbook(current_workbook_path, read_only=True, data_only=True)
                except:
                    # ファイルが開けない場合はスキップ
                    workbook = None

            # ブックが開けなかった場合はスキップ
            if workbook is None:
                progress_cnt += 1
                if progress:
                    progress.update(current=progress_cnt, status_msg=f'Processing: {progress_cnt}/{progress_max}')
                continue

            # シートを指定する
            worksheet = workbook[output_row['Sheet']]

            # キーワードカウント用辞書を初期化
            keyword_counts = {keyword:0 for keyword in keywords}
            # 既存のキーワードカウントを取得
            for keyword in keywords:
                val = self._pd_keyword.at[output_index, keyword]
                if val is not None and val != "":
                    keyword_counts[keyword] = val

            # キーワードがシート内に含まれているかチェックする
            for row in worksheet.iter_rows(values_only=True):
                # セルごとにチェック
                for cell in row:
                    if cell is None:
                        continue
                    cell_str = str(cell)
                    # キーワードごとにチェック
                    for keyword in keywords:
                        if keyword in cell_str:
                            keyword_counts[keyword] += 1

            # キーワードカウントをDataFrameに設定
            for keyword, count in keyword_counts.items():
                self._pd_keyword.at[output_index, keyword] = count

            # 進捗表示
            progress_cnt += 1
            if progress:
                progress.update(current=progress_cnt, status_msg=f'Processing: {progress_cnt}/{progress_max}')

        # 0を空文字に置換
        self._pd_keyword = self._pd_keyword.replace(0, "")

        # workbookを閉じる
        if workbook is not None:
            workbook.close()

        # 進捗表示(100%)
        if progress:
            progress.update(current=progress_max, status_msg='Completed')
    
    def _search_keyword_shape(self, keywords:list) -> None:
        """キーワード検索

        図形内のテキストにキーワードが含まれる箇所をカウントする

        Args:
            keywords (list): 検索キーワードリスト
        """
        # Excelアプリケーションを起動
        excel = win32com.client.Dispatch('Excel.Application')
        excel.Visible = False
        excel.DisplayAlerts = False

        # 進捗データ初期化
        progress_cnt = 0
        progress_max = self._pd_keyword.shape[0]
        workbook = None
        current_workbook_path = None
        # 進捗表示用フラグを初期化
        progress = CommonProgress(total=progress_max, task_msg=self._doc_type+' Keyword in Shapes') if self._enable_progress else None

        try:
            # ブック＋シートでキーワードを検索する
            for output_index, output_row in self._pd_keyword.iterrows():
                try:
                    # Bad Fileの場合はスキップ
                    if output_row['Sheet'] == "Bad File Error":
                        progress_cnt += 1
                        if progress:
                            progress.update(current=progress_cnt, status_msg=f'Processing: {progress_cnt}/{progress_max}')
                        continue

                    # フルパスを生成してOpen済みか確認してブックを開く 
                    full_workbook_path = os.path.abspath(os.path.join(output_row['Path'], output_row['Book']))
                    # 既に開いているブックと異なる場合は新たに開く
                    if(full_workbook_path != current_workbook_path):
                        # 既に開いているブックがあれば閉じる
                        if current_workbook_path is not None:
                            workbook.Close(SaveChanges=False)
                        current_workbook_path = full_workbook_path
                        try:
                            # ブックを開く
                            workbook = excel.Workbooks.Open(full_workbook_path, ReadOnly=True)
                        except:
                            # ファイルが開けない場合はスキップ
                            workbook = None

                    # ブックが開けなかった場合はスキップ
                    if workbook is None:
                        progress_cnt += 1
                        if progress:
                            progress.update(current=progress_cnt, status_msg=f'Processing: {progress_cnt}/{progress_max}')
                        continue

                    # シートを指定する
                    worksheet = workbook.Sheets(output_row['Sheet'])

                    # キーワードがシート内に含まれているかチェックする
                    for keyword in keywords:
                        # キーワード出現回数を初期化
                        count = 0 if self._pd_keyword.at[output_index, keyword] in (None, "") else self._pd_keyword.at[output_index, keyword]
                        # シート内の図形のテキストをチェック
                        for shape in worksheet.Shapes:
                            # グループ化された図形も再帰的にチェック
                            try:
                                count += self._search_keyword_shape_group(shape, keyword)
                            except:
                                pass
                        self._pd_keyword.at[output_index, keyword] = count
                except:
                    pass

                # 進捗表示
                progress_cnt += 1
                if progress:
                    progress.update(current=progress_cnt, status_msg=f'Processing: {progress_cnt}/{progress_max}')
        finally:
            # 0を空文字に置換
            self._pd_keyword = self._pd_keyword.replace(0, "")

            # Excelアプリケーションを終了
            try:
                if workbook is not None:
                    workbook.Close(SaveChanges=False)
            except:
                pass
            try:
                excel.Quit()
            except:
                pass
            del excel
            # 進捗表示(100%)
            if progress:
                progress.update(current=progress_max, status_msg='Completed')

    def _search_keyword_shape_group(self, shape, keyword) -> int:
            """グループ化された図形を再帰的にチェックするサブメソッド
            Args:
                shape: 図形オブジェクト
                keyword (str): 検索キーワード
            Returns:
                int: キーワード出現回数
            """
            count = 0
            # Type 6 は msoGroup (グループ化された図形)
            if shape.Type == 6:
                try:
                    # グループ内の図形を再帰的にチェック
                    for sub_shape in shape.GroupItems:
                        count += self._search_keyword_shape_group(sub_shape, keyword)
                except:
                    pass
            else:
                try:
                    # テキストを持っているか判定
                    if shape.HasTextFrame:
                        # テキストを取得してキーワードが含まれるかチェック
                        txt = shape.TextFrame.Characters().Text
                        if txt and keyword in txt:
                            count += 1
                except:
                    pass
            # キーワード出現回数を返す
            return count