from .abstract_search_docs import AbstractSearchDocs
import os
import pandas as pd
import openpyxl
import win32com.client
import warnings
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

class BaseSearchExcel(AbstractSearchDocs):
    """EXCEL検索基底抽象クラス
    """
    _is_search_shapes : bool = False # 図形内テキスト検索フラグ
    excel_extension = ['.xlsx', '.xlsm', '.xls'] # Excelファイルの拡張子リスト

    def __init__(self) -> None:
        """コンストラクタ
        """
        super().__init__()

    def __del__(self) -> None:
        """デストラクタ
        """
        pass

    def search_target(self, target_path:str, progress:bool=True) -> bool:
        """対象検索処理
        Args:
            target_path (str): 検索対象パス
            progress (bool): 進捗表示フラグ

        Returns:
            bool: True:成功, False:失敗
        """
        # excelファイルリストを取得
        excel_files = []
        for root, dirs, files in os.walk(target_path):
            for file in files:
                # 拡張子がExcelファイルの場合
                if file.endswith(tuple(self.excel_extension)):
                    excel_files.append(os.path.join(root, file))

        # exelファイルがある場合はシート名リストを取得
        if len(excel_files) > 0:
            self._search_sheet_list(excel_files, progress)
            return True
        else:
            return False

    def search_keyword(self, keywords:str, progress:bool=True) -> bool:
        """キーワード検索処理
        Args:
            keywords (str): 検索キーワード
            progress (bool): 進捗表示フラグ

        Returns:
            bool: True:成功, False:失敗
        """
        # キーワードがない場合は終了
        if keywords is None or len(keywords) == 0:
            return False
        # 対象検索結果がない場合は終了
        if self._pd_target is None or self._pd_target.empty:
            return False
        
        # キーワード検索結果を対象検索結果で初期化して、キーワードのカラムを追加
        self._pd_keyword = self._pd_target.copy(deep=True)
        for keyword in keywords:
            self._pd_keyword[keyword] = None

        # CELL内テキスト検索を実行
        self._search_keyword_cell(keywords)
        # 図形内テキスト検索を実行
        if self._is_search_shapes:
            self._search_keyword_shape(keywords)
    
    def enable_search_shapes(self, enable:bool=True) -> None:
        """図形内テキスト検索の有効/無効設定
        Args:
            enable (bool): True:有効, False:無効
        """
        self._is_search_shapes = enable

    #
    # protected methods
    #
    def _search_sheet_list(self, files:list, is_progress:bool=True) -> None:
        """
        シート名リスト取得
        Args:
            files (list): excelファイルのリスト（フルパス）
            progress (bool): 進捗表示フラグ
        """
        # Sheet名リストを取得
        progress_max = len(files)
        rows = []
        workbook = None

        for i, file in enumerate(files, 1):
            file_name = os.path.basename(file)
            file_path = os.path.dirname(file)
            try:
                # ブックを開く
                workbook = openpyxl.load_workbook(file, read_only=True, data_only=True)
                sheetnames = workbook.sheetnames
                # ブック名とシート名をリストに登録
                for sheetname in sheetnames:
                    rows.append({
                        'Path':file_path,
                        'Book':file_name,
                        'Sheet':sheetname
                    })
            except :
                rows.append({
                    'Path':file_path,
                    'Book':file_name,
                    'Sheet':"Bad File Error"
                })
            finally:
                if workbook is not None:
                    workbook.close()
                    workbook = None
            # 進捗表示
            if is_progress:
                percent = round((i / progress_max) * 100)
                self.show_progress(percent, 'Sheets検索', f'{i}/{progress_max}')
        # 列名を設定
        self._pd_target = pd.DataFrame(rows, columns=['Path','Book','Sheet'])
        # 0を空文字に置換
        output_df = self._pd_target.replace(0, "")
        self._pd_target = output_df

        # 進捗表示(100%)
        if is_progress and percent < 100:
            self.show_progress(100, 'Sheets検索', '完了しました')
    
    def _search_keyword_cell(self, keywords:list, is_progress:bool=True) -> None:
        """キーワード検索処理
        Args:
            keywords (list): 検索キーワードリスト
        """
        # 進捗データ初期化
        progress = 0
        progress_max = self._pd_keyword.shape[0]

        # ブック＋シートでキーワードを検索する
        workbook= None
        current_workbook_path = None

        for output_index, output_row in self._pd_keyword.iterrows():
            # Bad Fileの場合はスキップ
            if output_row['Sheet'] == "Bad File Error":
                progress += 1
                continue
            
            # フルパスを生成してOpen済みか確認
            full_workbook_path = os.path.join(output_row['Path'], output_row['Book'])
            if current_workbook_path != full_workbook_path:
                if workbook is not None:
                    workbook.close()

                current_workbook_path = full_workbook_path
                try:
                    # ブックを開く
                    workbook = openpyxl.load_workbook(current_workbook_path, read_only=True, data_only=True)
                except:
                    # ファイルが開けない場合はスキップ
                    workbook = None
            if workbook is None:
                progress += 1
                continue

            # シートを指定する
            worksheet = workbook[output_row['Sheet']]

            # キーワードカウント用辞書を初期化
            keyword_counts = {keyword:0 for keyword in keywords}
            for keyword in keywords:
                val = self._pd_keyword.at[output_index, keyword]
                if val is not None and val != "":
                    keyword_counts[keyword] = val
            # キーワードがシート内に含まれているかチェックする
            for row in worksheet.iter_rows(values_only=True):
                for cell in row:
                    if cell is None:
                        continue
                    cell_str = str(cell)
                    for keyword in keywords:
                        if keyword in cell_str:
                            keyword_counts[keyword] += 1
            # キーワードカウントをDataFrameに設定
            for keyword, count in keyword_counts.items():
                self._pd_keyword.at[output_index, keyword] = count

            # 進捗表示
            if is_progress:
                progress += 1
                percent = round((progress / progress_max) * 100)
                self.show_progress(percent, 'セル内キーワード検索', f'{progress}/{progress_max}')

        # 0を空文字に置換
        output_df = self._pd_keyword.replace(0, "")
        self._pd_keyword = output_df

        # workbookを閉じる
        if workbook is not None:
            workbook.close()

        # 進捗表示(100%)
        if is_progress and percent < 100:
            self.show_progress(100, 'セル内キーワード検索', '完了しました')
    
    def _search_keyword_shape(self, keywords:list, is_progress:bool=True) -> None:
        """キーワード検索
        図形内のテキストにキーワードが含まれる箇所をカウントする

        Args:
            keywords (list): 検索キーワードリスト
        """
        # Excelアプリケーションを起動
        excel = win32com.client.Dispatch('Excel.Application')
        excel.Visible = False
        excel.DisplayAlerts = False

        # 進捗データ初期化
        progress = 0
        progress_max = self._pd_keyword.shape[0]
        # ブック＋シートでキーワードを検索する
        workbook = None
        current_workbook_path = None

        try:
            for output_index, output_row in self._pd_keyword.iterrows():
                try:
                    # Bad Fileの場合はスキップ
                    if output_row['Sheet'] == "Bad File Error":
                        continue

                    full_workbook_path = os.path.abspath(os.path.join(output_row['Path'], output_row['Book']))
                    # ブックを開く（一度開いている場合はスキップする)
                    if(full_workbook_path != current_workbook_path):
                        if current_workbook_path is not None:
                            workbook.Close(SaveChanges=False)
                        current_workbook_path = full_workbook_path
                        try:
                            # ブックを開く
                            workbook = excel.Workbooks.Open(full_workbook_path, ReadOnly=True)
                        except:
                            # ファイルが開けない場合はスキップ
                            workbook = None
                    if workbook is None:
                        progress += 1
                        continue

                    # シートを指定する
                    worksheet = workbook.Sheets(output_row['Sheet'])

                    # キーワードがシート内に含まれているかチェックする
                    for keyword in keywords:
                        count = 0 if self._pd_keyword.at[output_index, keyword] in (None, "") else self._pd_keyword.at[output_index, keyword]
                        # シート内の図形のテキストをチェック
                        for shape in worksheet.Shapes:
                            # if shape.HasTextFrame:
                            try:
                                count += self._search_keyword_shape_group(shape, keyword)
                            except:
                                pass
                        self._pd_keyword.at[output_index, keyword] = count
                except:
                    pass

                # 進捗表示
                if is_progress:
                    progress += 1
                    percent = round((progress / progress_max) * 100)
                    self.show_progress(percent, '図形内キーワード検索', f'{progress}/{progress_max}')
        finally:
            # 0を空文字に置換
            output_df = self._pd_keyword.replace(0, "")
            self._pd_keyword = output_df

            # Excelアプリケーションを終了
            try:
                if workbook is not None:
                    workbook.Close(SaveChanges=False)
            except:
                pass
            try:
                excel.Quit()
            except:
                pass
            del excel
            # 進捗表示(100%)
            if is_progress and percent < 100:
                self.show_progress(100, '図形内キーワード検索', '完了しました')

    def _search_keyword_shape_group(self, shape, keyword) -> int:
            """グループ化された図形を再帰的にチェックするサブメソッド"""
            count = 0
            # Type 6 は msoGroup (グループ化された図形)
            if shape.Type == 6:
                try:
                    for sub_shape in shape.GroupItems:
                        count += self._search_keyword_shape_group(sub_shape, keyword)
                except:
                    pass
            else:
                try:
                    # テキストを持っているか判定
                    if shape.HasTextFrame:
                        txt = shape.TextFrame.Characters().Text
                        if txt and keyword in txt:
                            count += 1
                except:
                    pass
            return count